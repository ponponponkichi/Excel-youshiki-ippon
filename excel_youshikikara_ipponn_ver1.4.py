import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, Listbox
import pandas as pd
import openpyxl
import glob
import datetime
import os
import threading
import re
import ctypes

# --------------------------------------------------------------------
# ■ 高DPI対応設定
# --------------------------------------------------------------------
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

# --------------------------------------------------------------------
# ■ カスタムTkinterの設定
# --------------------------------------------------------------------
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("dark-blue")
ctk.set_widget_scaling(1.0)

# --------------------------------------------------------------------
# ■ メイン処理ロジック
# --------------------------------------------------------------------
def extract_cell_values(folder_path, sheet_config, cell_list, app_instance):
    """
    sheet_config: {'mode': 'name' or 'index', 'value': 'Sheet1' or 1}
    cell_list: ['A1', 'B2', ...]
    """
    try:
        app_instance.update_status("処理を開始します...", "processing")

        xlsx_files_path = glob.glob(os.path.join(folder_path, "*.xlsx"))
        xls_files_path = glob.glob(os.path.join(folder_path, "*.xls"))
        all_files_path = xlsx_files_path + xls_files_path
        
        if not all_files_path:
            raise ValueError("指定されたフォルダにExcelファイル(.xlsx／.xls)が見つかりません。")

        all_data = []
        processed_files_count = 0

        for file_path in all_files_path:
            try:
                row_data = {'ファイル名': os.path.basename(file_path)}
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                target_sheet = None
                target_sheet_name = ""

                # --- シートの特定処理 ---
                if sheet_config['mode'] == 'name':
                    sheet_name = sheet_config['value']
                    if sheet_name in workbook.sheetnames:
                        target_sheet = workbook[sheet_name]
                        target_sheet_name = sheet_name
                    else:
                        target_sheet = None # シートが見つからない
                
                elif sheet_config['mode'] == 'index':
                    # ユーザー入力は1始まり、プログラムは0始まり
                    idx = int(sheet_config['value']) - 1
                    if 0 <= idx < len(workbook.sheetnames):
                        target_sheet = workbook.worksheets[idx]
                        target_sheet_name = target_sheet.title
                    else:
                        target_sheet = None # インデックス範囲外

                # --- セル値の抽出処理 ---
                for cell_address in cell_list:
                    column_header = cell_address  # 列名はシンプルにセル番地とする

                    if target_sheet:
                        try:
                            # 指定セルが存在するか確認（結合セルなどの場合も考慮して取得）
                            val = target_sheet[cell_address].value
                            row_data[column_header] = val
                        except Exception:
                             row_data[column_header] = "セル無効"
                    else:
                        if sheet_config['mode'] == 'name':
                            row_data[column_header] = "シート名無し"
                        else:
                            row_data[column_header] = "シート順無し"
                
                # デバッグ用や確認用に取得した実際のシート名を記録したい場合は以下のような列を追加しても良い
                # row_data['_取得シート名'] = target_sheet_name

                all_data.append(row_data)
                processed_files_count += 1
                app_instance.update_status(f"処理中: {os.path.basename(file_path)}", "processing")

            except Exception as e:
                print(f"エラー: {os.path.basename(file_path)} の処理中にエラー: {e}")
                error_row = {'ファイル名': f"{os.path.basename(file_path)} (読込エラー)"}
                all_data.append(error_row)

        if not all_data:
            raise ValueError("有効なデータをどのファイルからも抽出できませんでした。")

        # --- ファイルの保存処理 ---
        df = pd.DataFrame(all_data)
        
        # 列の並び順を整理
        column_order = ['ファイル名'] + cell_list
        df = df.reindex(columns=column_order)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"セル抽出結果_{timestamp}.xlsx"
        output_file_path = os.path.join(folder_path, file_name)

        with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='抽出結果', index=False)

        success_message = (
            f"処理が完了しました！\n"
            f"{processed_files_count}個のファイルを処理しました。\n\n"
            f"保存先: {output_file_path}"
        )
        messagebox.showinfo("成功", success_message)
        app_instance.update_status("待機中...", "normal")

    except Exception as e:
        messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")
        app_instance.update_status("エラーが発生しました", "error")

# --------------------------------------------------------------------
# ■ アプリケーションクラス
# --------------------------------------------------------------------
class CellExtractorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # ウィンドウ設定
        self.title("Excel所定様式からデータセット一本_ver1.4")
        self.geometry("600x850")
        
        # フォント定義
        self.main_font = ("Yu Gothic UI", 14)
        self.bold_font = ("Yu Gothic UI", 14, "bold")
        self.small_font = ("Yu Gothic UI", 12)
        self.btn_font = ("Yu Gothic UI", 16, "bold")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # メインスクロールフレーム
        self.main_frame = ctk.CTkScrollableFrame(self, corner_radius=0)
        self.main_frame.grid(row=0, column=0, sticky="nsew")
        self.main_frame.grid_columnconfigure(0, weight=1)

        # UI構築
        self.create_folder_section()
        self.create_sheet_section() # 新設: Step 2
        self.create_cell_section()  # 変更: Step 3
        self.create_action_section()
        self.create_info_section()

    def create_folder_section(self):
        frame = ctk.CTkFrame(self.main_frame)
        frame.grid(row=0, column=0, padx=20, pady=10, sticky="ew")
        frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(frame, text="1. Excelファイルを入れた【フォルダ】を選択", font=self.bold_font).grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 5))

        self.folder_path_entry = ctk.CTkEntry(frame, placeholder_text="フォルダパス", font=self.main_font)
        self.folder_path_entry.grid(row=1, column=0, columnspan=2, sticky="ew", padx=(10, 5), pady=10)

        browse_btn = ctk.CTkButton(frame, text="参照", width=80, command=self.select_folder, font=self.main_font)
        browse_btn.grid(row=1, column=2, sticky="e", padx=(0, 10), pady=10)

    def create_sheet_section(self):
        """Step 2: シート指定（単一指定、ラジオボタン切り替え）"""
        frame = ctk.CTkFrame(self.main_frame)
        frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(frame, text="2. 抽出するセルのあるシートを指定", font=self.bold_font).grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 5))

        # ラジオボタン用の変数 (value: "name" or "index")
        self.sheet_mode_var = tk.StringVar(value="name")

        # --- オプションA: シート名で指定 ---
        rb_name = ctk.CTkRadioButton(
            frame, text="シート名で指定", variable=self.sheet_mode_var, value="name",
            font=self.main_font, command=self.toggle_sheet_inputs
        )
        rb_name.grid(row=1, column=0, sticky="w", padx=20, pady=5)

        self.sheet_name_entry = ctk.CTkEntry(frame, placeholder_text="例: Sheet1", font=self.main_font)
        self.sheet_name_entry.grid(row=1, column=1, columnspan=2, sticky="ew", padx=(0, 20), pady=5)

        # --- オプションB: 左からの順番で指定 ---
        rb_index = ctk.CTkRadioButton(
            frame, text="シートの左からの順番で指定", variable=self.sheet_mode_var, value="index",
            font=self.main_font, command=self.toggle_sheet_inputs
        )
        rb_index.grid(row=2, column=0, sticky="w", padx=20, pady=5)

        self.sheet_index_entry = ctk.CTkEntry(frame, placeholder_text="例: 1 ", font=self.main_font)
        self.sheet_index_entry.grid(row=2, column=1, columnspan=2, sticky="ew", padx=(0, 20), pady=5)

        # 初期状態のUI制御
        self.toggle_sheet_inputs()

    def create_cell_section(self):
        """Step 3: セル番地の指定（複数可、リスト管理）"""
        frame = ctk.CTkFrame(self.main_frame)
        frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(frame, text="3. セルの指定（複数選択可）", font=self.bold_font).grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(10, 5))

        # 入力エリア
        input_frame = ctk.CTkFrame(frame, fg_color="transparent")
        input_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        
        ctk.CTkLabel(input_frame, text="セル番地:", font=self.main_font).pack(side="left", padx=(10, 5))
        self.cell_entry = ctk.CTkEntry(input_frame, width=120, placeholder_text="例: A1", font=self.main_font)
        self.cell_entry.pack(side="left", padx=5)

        add_btn = ctk.CTkButton(input_frame, text="追加", width=80, command=self.add_cell_spec, font=self.main_font)
        add_btn.pack(side="left", padx=10)
        
        # Enterキーでも追加できるようにバインド
        self.cell_entry.bind('<Return>', lambda event: self.add_cell_spec())

        # リストボックスエリア
        list_container = ctk.CTkFrame(frame)
        list_container.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        frame.grid_rowconfigure(2, weight=1)
        list_container.grid_columnconfigure(0, weight=1)

        self.spec_listbox = Listbox(
            list_container, 
            height=8, 
            font=("Yu Gothic UI", 12),
            selectmode=tk.SINGLE, 
            activestyle="none", 
            borderwidth=0, 
            highlightthickness=0,
            bg="white",
            fg="black"
        )
        self.spec_listbox.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        scrollbar = tk.Scrollbar(list_container, orient="vertical", command=self.spec_listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.spec_listbox.config(yscrollcommand=scrollbar.set)

        # 操作ボタンエリア
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.grid(row=2, column=1, sticky="ns", padx=5, pady=10)
        
        ctk.CTkButton(btn_frame, text="▲ 上へ", width=60, command=self.move_item_up, font=self.small_font).pack(pady=2)
        ctk.CTkButton(btn_frame, text="▼ 下へ", width=60, command=self.move_item_down, font=self.small_font).pack(pady=2)
        ctk.CTkButton(btn_frame, text="削除", width=60, fg_color="#D32F2F", hover_color="#C62828", command=self.remove_cell_spec, font=self.small_font).pack(pady=(15, 2))

    def create_action_section(self):
        self.execute_button = ctk.CTkButton(
            self.main_frame, 
            text="処理を実行する", 
            command=self.start_processing,
            height=50,
            font=self.btn_font,
            fg_color="#107C41", 
            hover_color="#0C5C30"
        )
        self.execute_button.grid(row=3, column=0, padx=20, pady=(20, 10), sticky="ew")

        self.status_label = ctk.CTkLabel(self.main_frame, text="待機中...", font=self.small_font)
        self.status_label.grid(row=4, column=0, padx=20, pady=(0, 10))

    def create_info_section(self):
        info_text = (
            "【使用方法】\n"
            "複数Excelファイルから、指定した「同一のシート」内の「指定セル」の値を抽出します。\n\n"
            "1. 対象のExcelファイルが入ったフォルダを選択。\n"
            "2. 抽出対象のシートを指定（「シート名」または「左からの順番」）。\n"
            "3. 抽出したいセル番地（例: A1, B10）を追加。\n"
            "   ※リストの順序でExcelに出力されます。\n"
            "4. 「実行」ボタンで処理開始。結果はフォルダ内に保存されます。\n"
        )
        
        frame = ctk.CTkFrame(self.main_frame)
        frame.grid(row=5, column=0, padx=20, pady=(10, 30), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        label = ctk.CTkLabel(
            frame, 
            text=info_text, 
            justify="left", 
            anchor="w", 
            font=self.small_font
        )
        label.grid(row=0, column=0, padx=15, pady=15, sticky="w")

    # --- 機能メソッド ---

    def toggle_sheet_inputs(self):
        """ラジオボタンの選択に応じて入力欄の有効/無効を切り替える"""
        mode = self.sheet_mode_var.get()
        if mode == "name":
            self.sheet_name_entry.configure(state="normal", fg_color=["#F9F9FA", "#343638"]) # デフォルト色
            self.sheet_index_entry.configure(state="disabled", fg_color=["#EBEBEB", "#2B2B2B"]) # グレーアウト
        else:
            self.sheet_name_entry.configure(state="disabled", fg_color=["#EBEBEB", "#2B2B2B"])
            self.sheet_index_entry.configure(state="normal", fg_color=["#F9F9FA", "#343638"])

    def select_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.folder_path_entry.delete(0, tk.END)
            self.folder_path_entry.insert(0, folder_path)

    def add_cell_spec(self):
        cell = self.cell_entry.get().strip().upper()

        if not cell:
            return # 空の場合は何もしない
        
        if not re.match(r'^[A-Z]+[1-9][0-9]*$', cell):
            messagebox.showwarning("入力エラー", "セル番地の形式が正しくありません。\n例: A1, C10, AA25")
            return

        # 重複チェック
        current_list = self.spec_listbox.get(0, tk.END)
        if cell in current_list:
            messagebox.showinfo("情報", f"セル {cell} は既に追加されています。")
            self.cell_entry.delete(0, tk.END)
            return

        self.spec_listbox.insert(tk.END, cell)
        self.cell_entry.delete(0, tk.END)
        self.cell_entry.focus_set()

    def remove_cell_spec(self):
        selected_indices = self.spec_listbox.curselection()
        if not selected_indices:
            return
        for i in reversed(selected_indices):
            self.spec_listbox.delete(i)

    def move_item_up(self):
        selected_indices = self.spec_listbox.curselection()
        if not selected_indices:
            return
        idx = selected_indices[0]
        if idx > 0:
            item_text = self.spec_listbox.get(idx)
            self.spec_listbox.delete(idx)
            self.spec_listbox.insert(idx - 1, item_text)
            self.spec_listbox.select_set(idx - 1)

    def move_item_down(self):
        selected_indices = self.spec_listbox.curselection()
        if not selected_indices:
            return
        idx = selected_indices[0]
        if idx < self.spec_listbox.size() - 1:
            item_text = self.spec_listbox.get(idx)
            self.spec_listbox.delete(idx)
            self.spec_listbox.insert(idx + 1, item_text)
            self.spec_listbox.select_set(idx + 1)

    def update_status(self, message, status_type="normal"):
        color = "black"
        if ctk.get_appearance_mode() == "Dark":
            color = "white"
        
        if status_type == "error":
            color = "#ff5555"
        elif status_type == "processing":
            color = "#3b8ed0"
            
        self.status_label.configure(text=message, text_color=color)

    def start_processing(self):
        folder = self.folder_path_entry.get()
        sheet_mode = self.sheet_mode_var.get()
        
        # シート情報の取得と検証
        sheet_config = {'mode': sheet_mode, 'value': None}
        if sheet_mode == 'name':
            val = self.sheet_name_entry.get().strip()
            if not val:
                messagebox.showwarning("入力エラー", "シート名を入力してください。")
                return
            sheet_config['value'] = val
        else:
            val = self.sheet_index_entry.get().strip()
            if not val.isdigit() or int(val) < 1:
                messagebox.showwarning("入力エラー", "シート順序は1以上の整数で入力してください。")
                return
            sheet_config['value'] = int(val)

        # セルリストの取得
        cell_list = list(self.spec_listbox.get(0, tk.END))

        if not folder or not cell_list:
            messagebox.showwarning("入力エラー", "フォルダを選択し、セルを1つ以上指定してください。")
            return

        process_thread = threading.Thread(
            target=extract_cell_values, 
            args=(folder, sheet_config, cell_list, self)
        )
        process_thread.start()

if __name__ == "__main__":
    app = CellExtractorApp()
    app.mainloop()