import streamlit as st
import pandas as pd
import openpyxl
import datetime
import io
import re

# --------------------------------------------------------------------
# ■ ページ基本設定（縦一本のコードと100%同一）
# --------------------------------------------------------------------
st.set_page_config(
    page_title="Excel所定様式からデータセット一本 Web版",
    layout="centered"
)

# 大タイトル（上下の余白を極限までカット・縦一本のHTML構造を完全流用）
st.markdown("""
    <div style="margin-top: 5px; margin-bottom: 0px;">
        <h1 style="font-size: 2.0rem; color: var(--text-color); font-weight: 700; border: none; padding: 0; margin: 0;">
            Excel所定様式からデータセット一本
        </h1>
    </div>
""", unsafe_allow_html=True)
st.caption("複数Excelファイルから、指定した同一のシート内の指定セルの値を抽出します。")

# セッション状態の初期化
if "cell_list" not in st.session_state:
    st.session_state.cell_list = []

# --------------------------------------------------------------------
# ■ 1. ファイル選択セクション
# --------------------------------------------------------------------
st.markdown("""
    <div style="border-left: 5px solid #107C41; padding-left: 10px; margin-top: 25px; margin-bottom: 10px;">
        <h2 style="font-size: 1.3rem; color: var(--text-color); font-weight: 700; margin: 0; padding: 0; border: none;">
            1. Excelファイルを入れたフォルダを選択（複数ファイル選択）
        </h2>
    </div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "対象ファイルをすべて選択してください",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)


# --------------------------------------------------------------------
# ■ 2. シート指定セクション
# --------------------------------------------------------------------
st.markdown("""
    <div style="border-left: 5px solid #107C41; padding-left: 10px; margin-top: 25px; margin-bottom: 10px;">
        <h2 style="font-size: 1.3rem; color: var(--text-color); font-weight: 700; margin: 0; padding: 0; border: none;">
            2. 抽出するセルのあるシートを指定
        </h2>
    </div>
""", unsafe_allow_html=True)

sheet_mode = st.radio(
    "指定モード",
    options=["name", "index"],
    format_func=lambda x: "シート名で指定" if x == "name" else "シートの左からの順番で指定",
    horizontal=True,
    label_visibility="collapsed"
)

sheet_config = {'mode': sheet_mode, 'value': None}

if sheet_mode == "name":
    sheet_name_input = st.text_input("シート名を入力してください", placeholder="例: Sheet1")
    sheet_config['value'] = sheet_name_input.strip()
else:
    sheet_index_input = st.number_input("シートの順序（1始まりの整数）", min_value=1, value=1, step=1)
    sheet_config['value'] = int(sheet_index_input)


# --------------------------------------------------------------------
# ■ 3. セル指定セクション
# --------------------------------------------------------------------
st.markdown("""
    <div style="border-left: 5px solid #107C41; padding-left: 10px; margin-top: 25px; margin-bottom: 10px;">
        <h2 style="font-size: 1.3rem; color: var(--text-color); font-weight: 700; margin: 0; padding: 0; border: none;">
            3. セルの指定（複数選択可）
        </h2>
    </div>
""", unsafe_allow_html=True)

with st.form(key="cell_input_form", clear_on_submit=True):
    col_input, col_btn = st.columns([3, 1])
    with col_input:
        new_cell = st.text_input("セル番地", placeholder="例: A1", label_visibility="collapsed")
    with col_btn:
        add_pressed = st.form_submit_button("追加", use_container_width=True)

    if add_pressed and new_cell:
        formatted_cell = new_cell.strip().upper()
        if not re.match(r'^[A-Z]+[1-9][0-9]*$', formatted_cell):
            st.warning("入力エラー: セル番地の形式が正しくありません。（例: A1, C10）")
        elif formatted_cell in st.session_state.cell_list:
            st.info(f"情報: セル {formatted_cell} は既に追加されています。")
        else:
            st.session_state.cell_list.append(formatted_cell)

if st.session_state.cell_list:
    st.markdown("<div style='margin-top: 10px; margin-bottom: 5px; font-size: 0.9rem; font-weight: bold;'>指定済みセルリスト:</div>", unsafe_allow_html=True)
    
    to_remove = []
    for idx, cell in enumerate(st.session_state.cell_list):
        col_c, col_d = st.columns([5, 1])
        with col_c:
            st.markdown(f"<div style='font-size: 0.9rem; padding-left: 5px;'>列 {idx+1}: {cell}</div>", unsafe_allow_html=True)
        with col_d:
            if st.button("削除", key=f"del_{cell}_{idx}"):
                to_remove.append(cell)
    
    if to_remove:
        for c in to_remove:
            st.session_state.cell_list.remove(c)
        st.rerun()
else:
    st.caption("セル番地が指定されていません。")


# --------------------------------------------------------------------
# ■ 4. 処理の実行セクション
# --------------------------------------------------------------------
st.markdown("""
    <div style="border-left: 5px solid #107C41; padding-left: 10px; margin-top: 25px; margin-bottom: 10px;">
        <h2 style="font-size: 1.3rem; color: var(--text-color); font-weight: 700; margin: 0; padding: 0; border: none;">
            4. 処理の実行
        </h2>
    </div>
""", unsafe_allow_html=True)

# ボタンの深緑色スタイルを縦一本ツールから完全再現
st.markdown("""
    <style>
    div.stButton > button:first-child {
        background-color: #107C41 !important;
        color: white !important;
        border: none !important;
        height: 40px !important;
        font-size: 14px !important;
        font-weight: bold !important;
        border-radius: 4px !important;
    }
    div.stButton > button:first-child:hover {
        background-color: #0C5C30 !important;
    }
    </style>
""", unsafe_allow_html=True)

if st.button("処理を実行する", use_container_width=True):
    if not uploaded_files:
        st.error("入力エラー: ファイルを選択し、セルを1つ以上指定してください。")
    elif sheet_mode == "name" and not sheet_config['value']:
        st.error("入力エラー: シート名を入力してください。")
    elif not st.session_state.cell_list:
        st.error("入力エラー: 抽出対象のセル番地を指定してください。")
    else:
        try:
            with st.spinner("処理中..."):
                all_data = []
                processed_files_count = 0

                for file in uploaded_files:
                    try:
                        row_data = {'ファイル名': file.name}
                        workbook = openpyxl.load_workbook(file, data_only=True)
                        
                        target_sheet = None
                        if sheet_config['mode'] == 'name':
                            s_name = sheet_config['value']
                            if s_name in workbook.sheetnames:
                                target_sheet = workbook[s_name]
                        elif sheet_config['mode'] == 'index':
                            idx = sheet_config['value'] - 1
                            if 0 <= idx < len(workbook.sheetnames):
                                target_sheet = workbook.worksheets[idx]

                        for cell_address in st.session_state.cell_list:
                            column_header = cell_address
                            
                            if target_sheet:
                                try:
                                    val = target_sheet[cell_address].value
                                    row_data[column_header] = val
                                except Exception:
                                    row_data[column_header] = "セル無効"
                            else:
                                if sheet_config['mode'] == 'name':
                                    row_data[column_header] = "シート名無し"
                                else:
                                    row_data[column_header] = "シート順無し"

                        all_data.append(row_data)
                        processed_files_count += 1

                    except Exception:
                        error_row = {'ファイル名': f"{file.name} (読込エラー)"}
                        for cell_address in st.session_state.cell_list:
                            error_row[cell_address] = "エラー発生"
                        all_data.append(error_row)

                if all_data:
                    df = pd.DataFrame(all_data)
                    column_order = ['ファイル名'] + st.session_state.cell_list
                    df = df.reindex(columns=column_order)

                    st.success(f"処理完了！ {processed_files_count}個のファイルからデータを抽出しました。")
                    st.dataframe(df.head(10), use_container_width=True)

                    # Excel出力バッファ作成
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='抽出結果', index=False)
                    download_data = excel_buffer.getvalue()

                    now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"セル抽出結果_{now}.xlsx"

                    st.download_button(
                        label="統合データをダウンロードする",
                        data=download_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.error("エラー: 有効なデータを抽出できませんでした。")

        except Exception as e:
            st.error(f"処理中にエラーが発生しました:\\n{e}")


# --------------------------------------------------------------------
# ■ 説明インフォメーション（縦一本ツールの折りたたみ表記を完全再現）
# --------------------------------------------------------------------
st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
with st.expander("詳細な仕様・前提条件を確認する"):
    st.markdown("""
    **【使用方法】**
    * 複数Excelファイルから、指定した「同一のシート」内の「指定セル」の値を抽出します。
    * 1. 対象のExcelファイルが入ったフォルダから、ファイルを全選択してアップロードエリアにドロップします。
    * 2. 抽出対象のシートを指定します（「シート名」または「左からの順番」）。
    * 3. 抽出したいセル番地（例: A1, B10）を追加します。追加した順序でExcelの列に出力されます。
    * 4. 「処理を実行する」ボタンで処理を開始し、生成されたダウンロードボタンからファイルを保存してください。
    """)